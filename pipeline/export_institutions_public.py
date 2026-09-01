#!/usr/bin/env python3
"""서지 DB 의 **기관 정규화 표**만 공개용으로 내보낸다 (시트 한 장).

무엇을 내보내나 — 논문에 적힌 기관 문자열이 실제로 어느 기관이고, 그 위에 어떤
상위 조직이 있고, 어느 나라인지. ROR 정규화 + 운영자가 큐레이션한 Scopus 그룹
계층은 Crossref·OpenAlex 가 그대로 주지 않아 공개 가치가 있다.

무엇을 **안** 내보내나 — 저자·논문 단위 정보 전부. 원본 DB 에는 공개하면 안 되는
것이 세 종류 있고, 이 스크립트는 그 테이블을 아예 열지 않는다:

  1. `papers.metadata_json.pdf_path` — 운영자 Google Drive 절대경로. 계정
     이메일이 4,184행에 박혀 있다.
  2. `papers.header_raw` — PDF 앞머리 전체라 교신저자 이메일 2,074개가 나온다.
     논문에 적힌 공개 연락처지만, 긁어모아 기계가 읽는 한 파일로 재배포하는
     것은 논문에 실린 것과 전혀 다른 행위다.
  3. `papers.zotero_item_key` — 개인 Zotero 라이브러리 키.

`paper_institutions.raw_name` 은 읽되 그대로 싣지 않는다. 이 컬럼에도 이메일이
143건, 저자 나열이 섞여 있다(`Zhengwei Tao, Dingchu Zhang, … Alibaba Group`).
블랙리스트로 걸러내려는 시도는 두 방향 모두 실패했다 — 이름 나열 패턴은
`Oak Ridge National Laboratory,` 같은 실재 기관을 죽였고, 기관어 요구는
`Tencent`·`Genentech`·`DeepMind Technologies Limited` 를 죽였다. 그래서
**화이트리스트로 뒤집었다**: 명백히 안전할 때만 원문을 싣고, 아니면 그 칸을
비운다(기관·상위·국가는 그대로 남는다). 실측 커버리지 96%.

출력:
  reports/build/institutions_public.csv    — 정본. 텍스트라 diff·리뷰가 된다
  reports/build/institutions_public.xlsx   — 시트 한 장, 배포·열람용

Usage:
  python pipeline/export_institutions_public.py
"""

import argparse
import csv
import os
import re
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import PROJECT_ROOT

# ── raw_name 정리 ──
_EMAIL = re.compile(r"[\w.+-]+@[\w-]+\.[\w]{2,}")
_ORCID = re.compile(r"\b\d{4}-\d{4}-\d{4}-\d{3}[\dX]\b")
# "Correspondence to: 이름 <메일>" 같은 연락처 꼬리 전체
_CORRESP = re.compile(r"\bcorre\-?\s?spondence\s+to\s*:.*$", re.I)
# 소속 위첨자 마커가 이름 앞에 붙어 나온다: "23Max-Planck Institute…"
_LEAD_MARKER = re.compile(r"^[\d\s,;*†‡§¶]+")
_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# ── 안전 판정 ──
_ORG_WORD = re.compile(
    r"(univers|institut|istituto|college|school|laborator|\blab\b|cent(er|re)|academ|"
    r"hospital|department|\bdept\b|faculty|research|gmbh|\binc\b|\bltd\b|limited|"
    r"\bllc\b|corp|compan|foundation|ministr|agency|societ|associat|museum|observator|"
    r"clinic|division|consorti|council|bureau|polytech|ecole|hochschule|instituto|"
    r"universidad|technolog|science|engineering|medicine|health|genome|national)", re.I)
# "Firstname Lastname" 꼴. 전부 대문자인 약어(ETH, CMU)는 걸리지 않는다.
_NAMEISH = re.compile(r"\b[A-Z][a-z]{1,15}\s+[A-Z][a-z]{1,15}\b")
_PII = re.compile(r"[\w.+-]+@[\w-]+\.[\w]{2,}|\b\d{4}-\d{4}-\d{4}-\d{3}[\dX]\b")

MAX_RAW_CHARS = 160

COLUMNS = ["논문 기관명", "실제 기관명", "상위 기관명", "국가",
           "ROR ID", "논문 수", "근거", "검토 필요"]


def clean_raw_name(raw: str) -> str:
    """논문에 적힌 기관 문자열에서 개인정보·추출 잡음을 걷어낸다."""
    s = _CORRESP.sub("", raw or "")
    s = _EMAIL.sub("", s)
    s = _ORCID.sub("", s)
    s = _LEAD_MARKER.sub("", s)
    s = _CTRL.sub(" ", s)
    s = re.sub(r"[<>]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip(" ,;.·-")


def is_publishable_raw(s: str) -> bool:
    """이 문자열을 '논문 기관명' 으로 실어도 되는가 — 확실할 때만 True.

    기관어가 있으면 지명·학교명이 이름처럼 보이는 걸 감안해 이름 패턴을 5개까지
    허용하고(`Oak Ridge National Laboratory, Oak Ridge, TN, USA`), 기관어가 없으면
    회사명·약어로 보고 이름 패턴 1개·쉼표 1개까지만 허용한다(`NVIDIA`,
    `Google DeepMind`, `ETH Zurich`).
    """
    if not (4 <= len(s) <= MAX_RAW_CHARS):
        return False
    if _PII.search(s):
        return False
    names = len(_NAMEISH.findall(s))
    if _ORG_WORD.search(s):
        return names < 5
    return names <= 1 and s.count(",") <= 1


def fetch_rows(db_path):
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    # papers / authors / paper_authors / source_documents 는 열지 않는다.
    sql = """
        SELECT pi.raw_name, i.institution_name, i.parent_name,
               i.country_name_en, i.ror_id, pi.source,
               COUNT(DISTINCT pi.paper_id) AS n_papers,
               SUM(CASE WHEN pi.country_name <> '' AND i.country_name_en <> ''
                         AND pi.country_name <> i.country_name_en
                        THEN 1 ELSE 0 END) AS country_conflict
          FROM paper_institutions pi
          JOIN institutions i USING (institution_id)
         GROUP BY pi.raw_name, i.institution_id
         ORDER BY n_papers DESC, i.institution_name
    """
    rows, redacted = [], 0
    for raw, name, parent, country, ror, source, n, conflict in conn.execute(sql):
        cleaned = clean_raw_name(raw)
        if not is_publishable_raw(cleaned):
            cleaned = ""            # 기관·상위·국가는 남기고 원문만 비운다
            redacted += 1
        rows.append({
            "논문 기관명": cleaned,
            "실제 기관명": name,
            "상위 기관명": parent or "",
            "국가": country or "",
            "ROR ID": (ror or "").replace("https://ror.org/", ""),
            "논문 수": n,
            "근거": source,
            # 논문에 적힌 국가와 정규화된 기관의 국가가 어긋나면 대개 일반명
            # 오매칭이다 — "Georgia Tech" 가 조지아(국가)에 붙는 식.
            "검토 필요": "국가 불일치" if conflict else "",
        })
    conn.close()
    return rows, redacted


def write_csv(rows, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "기관 정규화"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    widths = {"논문 기관명": 58, "실제 기관명": 44, "상위 기관명": 32, "국가": 18,
              "ROR ID": 12, "논문 수": 8, "근거": 20, "검토 필요": 12}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[col]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def main():
    p = argparse.ArgumentParser(description="Export the public institution table")
    p.add_argument("--db", default=str(PROJECT_ROOT / ".cache" / "bibliography.sqlite3"))
    p.add_argument("--outdir", default=str(PROJECT_ROOT / "reports" / "build"))
    args = p.parse_args()

    rows, redacted = fetch_rows(args.db)
    csv_path = os.path.join(args.outdir, "institutions_public.csv")
    xlsx_path = os.path.join(args.outdir, "institutions_public.xlsx")
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)

    inst = len({r["실제 기관명"] for r in rows})
    parent = sum(1 for r in rows if r["상위 기관명"])
    ror = sum(1 for r in rows if r["ROR ID"])
    flagged = sum(1 for r in rows if r["검토 필요"])
    print(f"행 {len(rows):,}  (고유 기관 {inst:,})")
    print(f"  논문 기관명 실림 {len(rows)-redacted:,} ({100*(len(rows)-redacted)/len(rows):.0f}%)"
          f" · 안전하지 않아 비움 {redacted:,}")
    print(f"  상위 기관명 있음 {parent:,} ({100*parent/len(rows):.0f}%)")
    print(f"  ROR ID 있음      {ror:,} ({100*ror/len(rows):.0f}%)")
    print(f"  검토 필요        {flagged:,}")
    for path in (csv_path, xlsx_path):
        print(f"  {os.path.getsize(path)/1e6:6.2f}MB  {path}")


if __name__ == "__main__":
    main()
